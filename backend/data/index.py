"""
CRUD API: shipments, flights, equipment, auto_tasks, action_logs, directory import.
Действие передаётся через поле action в теле запроса.
"""
import json
import os
import base64
import io
import psycopg2
import openpyxl

SCHEMA = 't_p78311576_logistics_web_app_1'
CORS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, Authorization, X-Auth-Token, X-User-Id',
    'Access-Control-Max-Age': '86400',
}

def get_db():
    return psycopg2.connect(os.environ['DATABASE_URL'])

def q(val):
    if val is None:
        return 'NULL'
    return "'" + str(val).replace("'", "''") + "'"

def qn(val, default=0):
    try:
        return str(float(val)) if val not in (None, '') else str(default)
    except (ValueError, TypeError):
        return str(default)

def ok(data, status=200):
    return {'statusCode': status, 'headers': {**CORS, 'Content-Type': 'application/json'}, 'body': json.dumps(data, ensure_ascii=False, default=str)}

def err(msg, status=400):
    return {'statusCode': status, 'headers': {**CORS, 'Content-Type': 'application/json'}, 'body': json.dumps({'error': msg}, ensure_ascii=False)}

def get_session_user(conn, token):
    if not token:
        return None
    cur = conn.cursor()
    cur.execute(
        f"SELECT u.id, u.name, u.email, u.role FROM {SCHEMA}.sessions s "
        f"JOIN {SCHEMA}.users u ON u.id = s.user_id "
        f"WHERE s.token = {q(token)} AND s.expires_at > NOW()"
    )
    row = cur.fetchone()
    if not row:
        return None
    return {'id': str(row[0]), 'name': row[1], 'email': row[2], 'role': row[3]}

def row_to_shipment(r):
    return {
        'id': str(r[0]), 'number': r[1], 'request': r[2], 'client': r[3],
        'containerNumber': r[4], 'footage': r[5],
        'deliveryDate': str(r[6]) if r[6] else '', 'docsDate': str(r[7]) if r[7] else '',
        'inspectionDate': str(r[8]) if r[8] else '',
        'places': r[9], 'weight': float(r[10]) if r[10] else 0,
        'cargo': r[11], 'tempMode': r[12], 'vsdNumber': r[13],
        'status': r[14], 'shipmentType': r[15], 'terminal': r[16],
        'destination': r[17], 'gngCode': r[18], 'etsnvCode': r[19],
        'requestName': r[20], 'comment': r[21], 'dtNumber': r[22],
        'billOfLading': r[23], 'subsidy': r[24],
        'flightId': str(r[25]) if r[25] else '',
        'editedBy': r[26] or '', 'editedAt': str(r[27]) if r[27] else '',
    }

def row_to_flight(r):
    return {'id': str(r[0]), 'number': r[1], 'direction': r[2], 'planDate': str(r[3]) if r[3] else '', 'factDate': str(r[4]) if r[4] else '', 'status': r[5]}

def row_to_equipment(r):
    return {'id': str(r[0]), 'number': r[1], 'type': r[2], 'status': r[3], 'location': r[4], 'lastCheck': str(r[5]) if r[5] else '', 'comment': r[6] or '', 'size': r[7] or '40HC'}

def row_to_auto_task(r):
    return {'id': str(r[0]), 'type': r[1], 'date': str(r[2]) if r[2] else '', 'containerNumber': r[3], 'client': r[4], 'carrier': r[5], 'time': r[6], 'address': r[7], 'contact': r[8], 'terminalFrom': r[9], 'terminalTo': r[10], 'cargo': r[11], 'tempMode': r[12], 'status': r[13], 'comment': r[14] or '', 'krkNumber': r[15] or ''}

def row_to_log(r):
    return {'id': str(r[0]), 'userId': r[1], 'userName': r[2], 'action': r[3], 'entity': r[4], 'entityId': r[5], 'timestamp': str(r[6])}

def handler(event: dict, context) -> dict:
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS, 'body': ''}

    headers = event.get('headers') or {}
    token = headers.get('X-Auth-Token', '')

    body = {}
    if event.get('body'):
        try:
            body = json.loads(event['body'])
        except (ValueError, TypeError):
            pass

    action = body.get('action', '')
    b = body

    conn = get_db()
    try:
        user = get_session_user(conn, token)
        if not user:
            return err('Не авторизован', 401)

        cur = conn.cursor()

        # ── SHIPMENTS ──────────────────────────────────────────────────────────
        if action == 'get_shipments':
            cur.execute(
                f"SELECT id, number, request, client, container_number, footage, "
                f"delivery_date, docs_date, inspection_date, places, weight, cargo, temp_mode, "
                f"vsd_number, status, shipment_type, terminal, destination, gng_code, etsnv_code, "
                f"request_name, comment, dt_number, bill_of_lading, subsidy, flight_id, "
                f"edited_by, edited_at FROM {SCHEMA}.shipments ORDER BY created_at"
            )
            return ok({'shipments': [row_to_shipment(r) for r in cur.fetchall()]})

        if action == 'create_shipment':
            cur.execute(
                f"INSERT INTO {SCHEMA}.shipments "
                f"(number, request, client, container_number, footage, delivery_date, docs_date, "
                f"inspection_date, places, weight, cargo, temp_mode, vsd_number, status, "
                f"shipment_type, terminal, destination, gng_code, etsnv_code, request_name, "
                f"comment, dt_number, bill_of_lading, subsidy, flight_id, edited_by) VALUES ("
                f"{q(b.get('number',''))},{q(b.get('request',''))},{q(b.get('client',''))},"
                f"{q(b.get('containerNumber',''))},{q(b.get('footage',''))},"
                f"{q(b.get('deliveryDate') or None)},{q(b.get('docsDate') or None)},{q(b.get('inspectionDate') or None)},"
                f"{int(b.get('places',0))},{qn(b.get('weight',0))},"
                f"{q(b.get('cargo',''))},{q(b.get('tempMode',''))},{q(b.get('vsdNumber',''))},"
                f"{q(b.get('status','not_ready'))},{q(b.get('shipmentType','import'))},"
                f"{q(b.get('terminal',''))},{q(b.get('destination',''))},"
                f"{q(b.get('gngCode',''))},{q(b.get('etsnvCode',''))},{q(b.get('requestName',''))},"
                f"{q(b.get('comment',''))},{q(b.get('dtNumber',''))},{q(b.get('billOfLading',''))},"
                f"{q(b.get('subsidy',''))},{q(b.get('flightId') or None)},{q(b.get('editedBy',''))}"
                f") RETURNING id"
            )
            new_id = cur.fetchone()[0]
            conn.commit()
            return ok({'id': str(new_id)}, 201)

        if action == 'update_shipment':
            rid = b.get('id','')
            cur.execute(
                f"UPDATE {SCHEMA}.shipments SET "
                f"number={q(b.get('number',''))}, request={q(b.get('request',''))}, client={q(b.get('client',''))}, "
                f"container_number={q(b.get('containerNumber',''))}, footage={q(b.get('footage',''))}, "
                f"delivery_date={q(b.get('deliveryDate') or None)}, docs_date={q(b.get('docsDate') or None)}, "
                f"inspection_date={q(b.get('inspectionDate') or None)}, "
                f"places={int(b.get('places',0))}, weight={qn(b.get('weight',0))}, "
                f"cargo={q(b.get('cargo',''))}, temp_mode={q(b.get('tempMode',''))}, vsd_number={q(b.get('vsdNumber',''))}, "
                f"status={q(b.get('status','not_ready'))}, shipment_type={q(b.get('shipmentType','import'))}, "
                f"terminal={q(b.get('terminal',''))}, destination={q(b.get('destination',''))}, "
                f"gng_code={q(b.get('gngCode',''))}, etsnv_code={q(b.get('etsnvCode',''))}, request_name={q(b.get('requestName',''))}, "
                f"comment={q(b.get('comment',''))}, dt_number={q(b.get('dtNumber',''))}, bill_of_lading={q(b.get('billOfLading',''))}, "
                f"subsidy={q(b.get('subsidy',''))}, flight_id={q(b.get('flightId') or None)}, "
                f"edited_by={q(b.get('editedBy',''))}, edited_at=NOW(), updated_at=NOW() WHERE id={q(rid)}"
            )
            conn.commit()
            return ok({'ok': True})

        if action == 'delete_shipment':
            rid = b.get('id','')
            cur.execute(f"INSERT INTO {SCHEMA}.action_logs (user_id, user_name, action, entity, entity_id) VALUES ({q(user['id'])},{q(user['name'])},{q('Удаление')},{q('Отправка')},{q(rid)})")
            cur.execute(f"DELETE FROM {SCHEMA}.shipments WHERE id={q(rid)}")
            conn.commit()
            return ok({'ok': True})

        # ── FLIGHTS ────────────────────────────────────────────────────────────
        if action == 'get_flights':
            cur.execute(f"SELECT id, number, direction, plan_date, fact_date, status FROM {SCHEMA}.flights ORDER BY created_at")
            return ok({'flights': [row_to_flight(r) for r in cur.fetchall()]})

        if action == 'create_flight':
            cur.execute(
                f"INSERT INTO {SCHEMA}.flights (number, direction, plan_date, fact_date, status) "
                f"VALUES ({q(b.get('number',''))},{q(b.get('direction','moscow'))},{q(b.get('planDate') or None)},{q(b.get('factDate') or None)},{q(b.get('status','planned'))}) RETURNING id"
            )
            new_id = cur.fetchone()[0]
            conn.commit()
            return ok({'id': str(new_id)}, 201)

        if action == 'update_flight':
            rid = b.get('id','')
            cur.execute(
                f"UPDATE {SCHEMA}.flights SET number={q(b.get('number',''))}, direction={q(b.get('direction','moscow'))}, "
                f"plan_date={q(b.get('planDate') or None)}, fact_date={q(b.get('factDate') or None)}, "
                f"status={q(b.get('status','planned'))}, updated_at=NOW() WHERE id={q(rid)}"
            )
            conn.commit()
            return ok({'ok': True})

        # ── EQUIPMENT ──────────────────────────────────────────────────────────
        if action == 'get_equipment':
            cur.execute(f"SELECT id, number, type, status, location, last_check, comment, size FROM {SCHEMA}.equipment ORDER BY created_at")
            return ok({'equipment': [row_to_equipment(r) for r in cur.fetchall()]})

        if action == 'create_equipment':
            cur.execute(
                f"INSERT INTO {SCHEMA}.equipment (number, type, status, location, last_check, comment, size) "
                f"VALUES ({q(b.get('number',''))},{q(b.get('type','container'))},{q(b.get('status','unchecked'))},{q(b.get('location',''))},{q(b.get('lastCheck') or None)},{q(b.get('comment',''))},{q(b.get('size','40HC'))}) RETURNING id"
            )
            new_id = cur.fetchone()[0]
            conn.commit()
            return ok({'id': str(new_id)}, 201)

        if action == 'update_equipment':
            rid = b.get('id','')
            cur.execute(
                f"UPDATE {SCHEMA}.equipment SET number={q(b.get('number',''))}, type={q(b.get('type','container'))}, "
                f"status={q(b.get('status','unchecked'))}, location={q(b.get('location',''))}, last_check={q(b.get('lastCheck') or None)}, "
                f"comment={q(b.get('comment',''))}, size={q(b.get('size','40HC'))}, updated_at=NOW() WHERE id={q(rid)}"
            )
            cur.execute(f"INSERT INTO {SCHEMA}.action_logs (user_id, user_name, action, entity, entity_id) VALUES ({q(user['id'])},{q(user['name'])},{q('Редактирование оборудования')},{q('Оборудование')},{q(rid)})")
            conn.commit()
            return ok({'ok': True})

        if action == 'delete_equipment':
            rid = b.get('id','')
            cur.execute(f"DELETE FROM {SCHEMA}.equipment WHERE id={q(rid)}")
            conn.commit()
            return ok({'ok': True})

        # ── AUTO TASKS ─────────────────────────────────────────────────────────
        if action == 'get_auto_tasks':
            cur.execute(
                f"SELECT id, type, date, container_number, client, carrier, time, address, "
                f"contact, terminal_from, terminal_to, cargo, temp_mode, status, comment, krk_number "
                f"FROM {SCHEMA}.auto_tasks ORDER BY created_at"
            )
            return ok({'auto_tasks': [row_to_auto_task(r) for r in cur.fetchall()]})

        if action == 'create_auto_task':
            cur.execute(
                f"INSERT INTO {SCHEMA}.auto_tasks (type, date, container_number, client, carrier, time, address, contact, terminal_from, terminal_to, cargo, temp_mode, status, comment, krk_number) "
                f"VALUES ({q(b.get('type','movement'))},{q(b.get('date') or None)},{q(b.get('containerNumber',''))},{q(b.get('client',''))},{q(b.get('carrier',''))},{q(b.get('time',''))},{q(b.get('address',''))},{q(b.get('contact',''))},{q(b.get('terminalFrom',''))},{q(b.get('terminalTo',''))},{q(b.get('cargo',''))},{q(b.get('tempMode',''))},{q(b.get('status','planned'))},{q(b.get('comment',''))},{q(b.get('krkNumber',''))}) RETURNING id"
            )
            new_id = cur.fetchone()[0]
            conn.commit()
            return ok({'id': str(new_id)}, 201)

        if action == 'update_auto_task':
            rid = b.get('id','')
            cur.execute(
                f"UPDATE {SCHEMA}.auto_tasks SET type={q(b.get('type','movement'))}, date={q(b.get('date') or None)}, "
                f"container_number={q(b.get('containerNumber',''))}, client={q(b.get('client',''))}, carrier={q(b.get('carrier',''))}, "
                f"time={q(b.get('time',''))}, address={q(b.get('address',''))}, contact={q(b.get('contact',''))}, "
                f"terminal_from={q(b.get('terminalFrom',''))}, terminal_to={q(b.get('terminalTo',''))}, "
                f"cargo={q(b.get('cargo',''))}, temp_mode={q(b.get('tempMode',''))}, status={q(b.get('status','planned'))}, "
                f"comment={q(b.get('comment',''))}, krk_number={q(b.get('krkNumber',''))}, updated_at=NOW() WHERE id={q(rid)}"
            )
            conn.commit()
            return ok({'ok': True})

        if action == 'delete_auto_task':
            rid = b.get('id','')
            cur.execute(f"DELETE FROM {SCHEMA}.auto_tasks WHERE id={q(rid)}")
            conn.commit()
            return ok({'ok': True})

        # ── ACTION LOGS ────────────────────────────────────────────────────────
        if action == 'get_action_logs':
            cur.execute(f"SELECT id, user_id, user_name, action, entity, entity_id, timestamp FROM {SCHEMA}.action_logs ORDER BY timestamp DESC LIMIT 500")
            return ok({'action_logs': [row_to_log(r) for r in cur.fetchall()]})

        if action == 'create_action_log':
            cur.execute(
                f"INSERT INTO {SCHEMA}.action_logs (user_id, user_name, action, entity, entity_id) "
                f"VALUES ({q(b.get('userId',''))},{q(b.get('userName',''))},{q(b.get('action',''))},{q(b.get('entity',''))},{q(b.get('entityId',''))}) RETURNING id"
            )
            new_id = cur.fetchone()[0]
            conn.commit()
            return ok({'id': str(new_id)}, 201)

        # ── DATABASE DIRECTORY ─────────────────────────────────────────────────

        # Универсальный CRUD для справочников
        DIR_TABLES = {
            'clients': ('db_clients', ['name', 'inn', 'contact_person', 'phone', 'email', 'comment']),
            'contractors': ('db_contractors', ['name', 'type', 'inn', 'contact_person', 'phone', 'email', 'comment']),
            'containers': ('db_containers', ['number', 'size', 'owner', 'comment']),
            'vehicles': ('db_vehicles', ['plate', 'driver_name', 'driver_phone', 'carrier', 'comment']),
            'vessels': ('db_vessels', ['name', 'flag', 'imo', 'comment']),
            'wagons': ('db_wagons', ['number', 'type', 'owner', 'comment']),
            'dgk': ('db_dgk', ['number', 'owner', 'comment']),
            'egk': ('db_egk', ['number', 'owner', 'comment']),
            'ndgu': ('db_ndgu', ['number', 'owner', 'comment']),
            'stations': ('db_stations', ['name', 'code', 'region', 'comment']),
            'terminals': ('db_terminals', ['name', 'city', 'address', 'contact', 'comment']),
            'cargo': ('db_cargo', ['name', 'gng_code', 'etsnv_code', 'temp_mode', 'comment']),
            'cities': ('db_cities', ['name', 'region', 'comment']),
        }

        for dir_key, (tbl, fields) in DIR_TABLES.items():
            if action == f'db_get_{dir_key}':
                cols = ', '.join(['id'] + fields + ['created_at'])
                cur.execute(f"SELECT {cols} FROM {SCHEMA}.{tbl} ORDER BY id")
                rows = []
                for r in cur.fetchall():
                    row = {'id': str(r[0])}
                    for i, f in enumerate(fields):
                        row[f] = r[i+1] or ''
                    row['created_at'] = str(r[len(fields)+1]) if r[len(fields)+1] else ''
                    rows.append(row)
                return ok({dir_key: rows})

            if action == f'db_create_{dir_key}':
                cols = ', '.join(fields)
                vals = ', '.join([q(b.get(f, '')) for f in fields])
                cur.execute(f"INSERT INTO {SCHEMA}.{tbl} ({cols}) VALUES ({vals}) RETURNING id")
                new_id = cur.fetchone()[0]
                conn.commit()
                return ok({'id': str(new_id)}, 201)

            if action == f'db_update_{dir_key}':
                rid = b.get('id', '')
                sets = ', '.join([f"{f}={q(b.get(f,''))}" for f in fields])
                cur.execute(f"UPDATE {SCHEMA}.{tbl} SET {sets}, updated_at=NOW() WHERE id={q(rid)}")
                conn.commit()
                return ok({'ok': True})

            if action == f'db_delete_{dir_key}':
                rid = b.get('id', '')
                cur.execute(f"DELETE FROM {SCHEMA}.{tbl} WHERE id={q(rid)}")
                conn.commit()
                return ok({'ok': True})

        # ── IMPORT DIRECTORY FROM EXCEL ────────────────────────────────────────
        if action == 'import_directory':
            dir_key = b.get('dir_key', '')
            file_b64 = b.get('file', '')
            if not dir_key or dir_key not in DIR_TABLES:
                return err('Неизвестный справочник', 400)
            if not file_b64:
                return err('Файл не передан', 400)

            tbl, fields = DIR_TABLES[dir_key]
            try:
                file_bytes = base64.b64decode(file_b64)
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
                ws = wb.active
            except Exception as ex:
                return err(f'Ошибка чтения Excel: {ex}', 400)

            # Первая строка — заголовки, маппинг по совпадению с именами полей (регистронезависимо)
            header_row = [str(c.value or '').strip().lower() for c in ws[1]]
            col_idx = {}
            for field in fields:
                for ci, h in enumerate(header_row):
                    if h == field.lower() or h == field.replace('_', ' ').lower():
                        col_idx[field] = ci
                        break

            inserted = 0
            skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                vals = {}
                for field in fields:
                    ci = col_idx.get(field)
                    vals[field] = str(row[ci]).strip() if ci is not None and row[ci] is not None else ''

                # Пропускаем строки без основного поля
                if not vals.get(fields[0], ''):
                    skipped += 1
                    continue

                cols_str = ', '.join(fields)
                vals_str = ', '.join([q(vals[f]) for f in fields])
                cur.execute(f"INSERT INTO {SCHEMA}.{tbl} ({cols_str}) VALUES ({vals_str})")
                inserted += 1

            conn.commit()
            return ok({'inserted': inserted, 'skipped': skipped})

        return err('Неизвестное действие', 400)
    finally:
        conn.close()